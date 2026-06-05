"""Tests unitaires pour l'import/export CSV et Excel (sans S3, sans serveur FastAPI)."""
import csv
import io
import uuid

from app.models import ClassificationEntry, ClassificationFile
from app.services.classification_service import ClassificationService


# ---------------------------------------------------------------------------
# Helpers (copié de import_export.py pour les tests autonomes)
# ---------------------------------------------------------------------------

CSV_HEADERS = ["code", "nom", "definition", "annotations", "parent_code"]


def _entries_to_rows(entries: list[ClassificationEntry]) -> list[list[str]]:
    rows = []
    for e in entries:
        rows.append([
            e.code,
            e.nom,
            e.definition,
            " ; ".join(e.annotations),
            e.parent_code or "",
        ])
    return rows


def _rows_to_entries(rows: list[dict]) -> list[ClassificationEntry]:
    entries = []
    for row in rows:
        annotations_raw = row.get("annotations", "").strip()
        annotations = [a.strip() for a in annotations_raw.split(";") if a.strip()] if annotations_raw else []
        parent_code = row.get("parent_code", "").strip() or None
        entries.append(ClassificationEntry(
            id=str(uuid.uuid4()),
            code=row["code"].strip(),
            nom=row["nom"].strip(),
            definition=row.get("definition", "").strip(),
            annotations=annotations,
            parent_code=parent_code,
        ))
    return entries


def _make_service_with_entries() -> ClassificationService:
    """Crée un ClassificationService en mémoire avec 2 entrées de test."""
    svc = ClassificationService()
    data = {
        "type": "test-type",
        "version": "1.0.0",
        "description": "Classification de test",
        "entries": [
            {
                "id": "id-1",
                "code": "ROOT_01",
                "nom": "Racine",
                "definition": "Définition racine",
                "annotations": ["annot1", "annot2"],
                "parent_code": None,
            },
            {
                "id": "id-2",
                "code": "ROOT_01_01",
                "nom": "Enfant",
                "definition": "Définition enfant",
                "annotations": ["annot3"],
                "parent_code": "ROOT_01",
            },
        ],
    }
    svc.load_from_dict(data)
    return svc


# ---------------------------------------------------------------------------
# Test 1 : export CSV — vérification des en-têtes et des valeurs
# ---------------------------------------------------------------------------

def test_export_csv_contenu():
    svc = _make_service_with_entries()
    cf = svc.get_by_type("test-type")
    assert cf is not None

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_HEADERS)
    for row in _entries_to_rows(cf.entries):
        writer.writerow(row)

    csv_content = output.getvalue()
    reader = csv.DictReader(io.StringIO(csv_content))
    rows = list(reader)

    # Vérification des en-têtes
    assert reader.fieldnames == CSV_HEADERS

    # Vérification de la première ligne (racine)
    assert rows[0]["code"] == "ROOT_01"
    assert rows[0]["nom"] == "Racine"
    assert rows[0]["definition"] == "Définition racine"
    assert rows[0]["annotations"] == "annot1 ; annot2"
    assert rows[0]["parent_code"] == ""

    # Vérification de la deuxième ligne (enfant)
    assert rows[1]["code"] == "ROOT_01_01"
    assert rows[1]["nom"] == "Enfant"
    assert rows[1]["parent_code"] == "ROOT_01"


# ---------------------------------------------------------------------------
# Test 2 : round-trip CSV (export puis import)
# ---------------------------------------------------------------------------

def test_import_csv_round_trip():
    svc = _make_service_with_entries()
    cf = svc.get_by_type("test-type")
    assert cf is not None

    # Export
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_HEADERS)
    for row in _entries_to_rows(cf.entries):
        writer.writerow(row)

    bom = "﻿"
    csv_bytes = (bom + output.getvalue()).encode("utf-8")

    # Re-import
    text = csv_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    entries = _rows_to_entries(rows)

    assert len(entries) == 2

    # Première entrée (racine)
    e0 = entries[0]
    assert e0.code == "ROOT_01"
    assert e0.nom == "Racine"
    assert e0.definition == "Définition racine"
    assert e0.annotations == ["annot1", "annot2"]
    assert e0.parent_code is None

    # Deuxième entrée (enfant)
    e1 = entries[1]
    assert e1.code == "ROOT_01_01"
    assert e1.nom == "Enfant"
    assert e1.annotations == ["annot3"]
    assert e1.parent_code == "ROOT_01"


# ---------------------------------------------------------------------------
# Test 3 : export Excel — en-têtes en gras et valeurs correctes
# ---------------------------------------------------------------------------

def test_export_excel_contenu():
    import openpyxl
    from openpyxl.styles import Font

    svc = _make_service_with_entries()
    cf = svc.get_by_type("test-type")
    assert cf is not None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "test-type"

    ws.append(CSV_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in _entries_to_rows(cf.entries):
        ws.append(row)

    # Largeurs automatiques
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Relire le fichier pour vérification
    wb2 = openpyxl.load_workbook(buf)
    ws2 = wb2.active

    # Vérification du nom de la feuille
    assert ws2.title == "test-type"

    # Vérification des en-têtes et du gras
    header_row = list(ws2.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    header_values = [cell.value for cell in header_row]
    assert header_values == CSV_HEADERS
    for cell in header_row:
        assert cell.font.bold, f"La cellule d'en-tête {cell.coordinate} doit être en gras"

    # Vérification des données (ligne 2)
    rows = list(ws2.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2

    r0 = rows[0]
    assert r0[0] == "ROOT_01"
    assert r0[1] == "Racine"
    assert r0[3] == "annot1 ; annot2"
    assert (r0[4] or "") == ""  # parent_code vide pour la racine (None ou "" selon openpyxl)

    r1 = rows[1]
    assert r1[0] == "ROOT_01_01"
    assert r1[4] == "ROOT_01"
