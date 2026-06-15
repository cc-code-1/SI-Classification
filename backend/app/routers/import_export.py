import csv
import io
import json
import os
import uuid
from fastapi import APIRouter, HTTPException, UploadFile, File, Query, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from app.models import ClassificationFile, ClassificationEntry
from app.services.classification_service import service
from app.services import normalizer
from app.auth import require_user

router = APIRouter(prefix="/api", tags=["import-export"], dependencies=[Depends(require_user)])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

CSV_HEADERS = ["code", "nom", "definition", "annotations", "parent_code"]


def _entries_to_rows(entries: list[ClassificationEntry]) -> list[list[str]]:
    rows = []
    for e in entries:
        rows.append([
            e.code,
            e.nom,
            e.definition,
            " ; ".join(e.annotations),
            e.parent_code or "",
        ])
    return rows


def _decode_csv_bytes(content: bytes) -> str:
    """Décode des octets CSV en tolérant les encodages courants.

    Excel (Windows, France) exporte souvent en CP1252/Latin-1, pas en UTF-8.
    On tente l'UTF-8 (avec BOM) puis on se rabat sur CP1252.
    """
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    # Dernier recours : on ignore les octets invalides plutôt que de planter.
    return content.decode("utf-8", errors="replace")


def _read_csv_rows(text: str) -> list[dict]:
    """Lit un CSV en détectant le séparateur (',' ou ';') et en normalisant
    les en-têtes (minuscules, sans espaces ni BOM résiduel)."""
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        # Heuristique de repli : si la première ligne contient plus de ';' que
        # de ',', on choisit ';' (cas typique des exports Excel français).
        first_line = sample.splitlines()[0] if sample.splitlines() else ""
        delimiter = ";" if first_line.count(";") > first_line.count(",") else ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for raw in reader:
        rows.append({
            (k or "").strip().lstrip("﻿").lower(): (v or "")
            for k, v in raw.items()
        })
    return rows


_COLUMN_ALIASES: dict[str, str] = {
    "sous-domaine": "nom",
    "sous_domaine": "nom",
    "domaine": "nom",
    "libellé": "nom",
    "libelle": "nom",
    "label": "nom",
    "définition": "definition",
    "definition": "definition",
    "defintion": "definition",
}


def _normalize_row(row: dict) -> dict:
    return {_COLUMN_ALIASES.get(k, k): v for k, v in row.items()}


def _rows_to_entries(rows: list[dict]) -> list[ClassificationEntry]:
    from app.services.normalizer import _derive_parent_from_code

    normalized = []
    for i, row in enumerate(rows, start=2):
        row = _normalize_row(row)
        if "code" not in row or "nom" not in row:
            raise ValueError(
                "Colonnes 'code' et 'nom' obligatoires. Colonnes trouvées : "
                + (", ".join(sorted(k for k in row.keys() if k)) or "(aucune)")
            )
        code = (row.get("code") or "").strip()
        if not code:
            raise ValueError(f"Ligne {i} : la colonne 'code' est vide.")
        normalized.append((i, row, code))

    all_codes = {code for _, _, code in normalized}
    entries = []
    for i, row, code in normalized:
        parent_code = (row.get("parent_code") or "").strip() or None
        # Si pas de colonne parent_code, on déduit la hiérarchie depuis la structure du code
        if not parent_code:
            parent_code = _derive_parent_from_code(code, all_codes)
        annotations_raw = (row.get("annotations") or "").strip()
        annotations = [a.strip() for a in annotations_raw.split(";") if a.strip()] if annotations_raw else []
        entries.append(ClassificationEntry(
            id=str(uuid.uuid4()),
            code=code,
            nom=(row.get("nom") or "").strip(),
            definition=(row.get("definition") or "").strip(),
            annotations=annotations,
            parent_code=parent_code,
        ))
    return entries


# ---------------------------------------------------------------------------
# JSON Import / Export (existing)
# ---------------------------------------------------------------------------

@router.post("/import", response_model=ClassificationFile)
async def import_classification(
    file: UploadFile = File(...),
    auto_convert: bool = Query(
        True,
        description="Convertit automatiquement les formats non canoniques "
        "(champ 'sous_domaine'/'domaine' au lieu de 'nom', hiérarchie déduite "
        "du code, liste d'entrées sans enveloppe, etc.).",
    ),
    type: str = Query(
        None,
        description="Surcharge le champ 'type' du fichier. Utile pour importer "
        "plusieurs fichiers ayant le même type sans qu'ils s'écrasent.",
    ),
):
    """
    Importe un fichier JSON de classification.

    Si `auto_convert` est activé (par défaut), le fichier est normalisé vers le
    schéma interne : libellé déduit de divers champs, parent déduit du code, etc.
    Si `type` est fourni, il remplace le champ 'type' du fichier après normalisation.
    """
    if not file.filename or not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format JSON (.json)")
    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"JSON invalide : {e}")

    if auto_convert:
        # Type de repli = nom du fichier sans extension (ex: "sous_domaine.json")
        fallback_type = os.path.splitext(file.filename)[0]
        try:
            data = normalizer.normalize(data, fallback_type=fallback_type)
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"Conversion impossible : {e}")

    if type:
        data["type"] = type

    try:
        cf = service.load_from_dict(data)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Structure du fichier invalide : {e}")
    return cf


@router.post("/import/preview")
async def preview_import(file: UploadFile = File(...), auto_convert: bool = Query(True)):
    """
    Analyse un fichier sans l'importer : retourne le type détecté, le nombre
    d'entrées, et un aperçu des 3 premières entrées normalisées. Permet à
    l'interface d'afficher un récapitulatif avant validation.
    """
    if not file.filename or not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format JSON (.json)")
    content = await file.read()
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=400, detail=f"JSON invalide : {e}")

    was_converted = auto_convert and not normalizer.is_canonical(data)
    if auto_convert:
        fallback_type = os.path.splitext(file.filename)[0]
        data = normalizer.normalize(data, fallback_type=fallback_type)

    entries = data.get("entries", [])
    return {
        "type": data.get("type"),
        "version": data.get("version"),
        "description": data.get("description"),
        "entry_count": len(entries),
        "was_converted": was_converted,
        "sample": entries[:3],
    }


@router.get("/export/{type_}")
def export_classification(
    type_: str,
    format: str = Query(
        "nested",
        description="Format d'export : 'nested' (imbriqué, lisible, recommandé) "
        "ou 'flat' (liste plate avec parent_code).",
    ),
):
    """Exporte le JSON complet d'un type de classification, à plat ou imbriqué."""
    if format == "flat":
        data = service.export_to_dict(type_)
    else:
        data = service.export_nested(type_)
    if data is None:
        raise HTTPException(status_code=404, detail=f"Type '{type_}' introuvable")
    suffix = "" if format == "nested" else "_plat"
    json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(json_bytes),
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{type_}{suffix}.json"',
        },
    )


# ---------------------------------------------------------------------------
# CSV Import / Export
# ---------------------------------------------------------------------------

@router.post("/import/csv", response_model=ClassificationFile)
async def import_csv(
    file: UploadFile = File(...),
    type: str = Query(..., description="Nom du type de classification (ex: sous-domaine)"),
):
    """Importe un fichier CSV de classification."""
    content = await file.read()
    text = _decode_csv_bytes(content)

    try:
        rows = _read_csv_rows(text)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Lecture du CSV impossible : {e}")

    if not rows:
        raise HTTPException(status_code=400, detail="Le fichier CSV est vide ou mal formé")

    try:
        entries = _rows_to_entries(rows)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erreur lors du parsing CSV : {e}")

    data = {
        "type": type,
        "version": "1.0.0",
        "description": "",
        "entries": [e.model_dump() for e in entries],
    }
    try:
        cf = service.load_from_dict(data)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Structure invalide : {e}")
    return cf


@router.get("/export/{type_}/csv")
def export_csv(type_: str):
    """Exporte la classification en CSV (UTF-8 avec BOM)."""
    cf = service.get_by_type(type_)
    if cf is None:
        raise HTTPException(status_code=404, detail=f"Type '{type_}' introuvable")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(CSV_HEADERS)
    for row in _entries_to_rows(cf.entries):
        writer.writerow(row)

    # UTF-8 BOM pour compatibilité Excel
    bom = "﻿"
    csv_bytes = (bom + output.getvalue()).encode("utf-8")

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{type_}.csv"'},
    )


# ---------------------------------------------------------------------------
# Excel Import / Export
# ---------------------------------------------------------------------------

@router.post("/import/excel", response_model=ClassificationFile)
async def import_excel(
    file: UploadFile = File(...),
    type: str = Query(..., description="Nom du type de classification (ex: sous-domaine)"),
):
    """Importe un fichier Excel (.xlsx) de classification."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier Excel invalide : {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    # En-têtes normalisés (minuscules, sans espaces) pour tolérer "Code", " nom ", etc.
    headers = [str(h).strip().lower() if h is not None else "" for h in next(rows_iter, [])]

    if not headers:
        raise HTTPException(status_code=400, detail="Le fichier Excel est vide")

    rows = []
    for row in rows_iter:
        row_dict = {headers[i]: (str(row[i]).strip() if row[i] is not None else "") for i in range(len(headers))}
        rows.append(row_dict)

    if not rows:
        raise HTTPException(status_code=400, detail="Aucune donnée dans le fichier Excel")

    try:
        entries = _rows_to_entries(rows)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erreur lors du parsing Excel : {e}")

    data = {
        "type": type,
        "version": "1.0.0",
        "description": "",
        "entries": [e.model_dump() for e in entries],
    }
    try:
        cf = service.load_from_dict(data)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Structure invalide : {e}")
    return cf


@router.get("/export/{type_}/xlsx")
def export_xlsx(type_: str):
    """Exporte la classification en Excel (.xlsx)."""
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font  # type: ignore
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")

    cf = service.get_by_type(type_)
    if cf is None:
        raise HTTPException(status_code=404, detail=f"Type '{type_}' introuvable")

    wb = openpyxl.Workbook()
    sheet_name = type_[:31]  # limite Excel
    ws = wb.active
    ws.title = sheet_name

    # En-têtes en gras
    ws.append(CSV_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in _entries_to_rows(cf.entries):
        ws.append(row)

    # Largeurs de colonnes automatiques
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{type_}.xlsx"'},
    )
